"""
Quantification pipeline for segmented intestinal ROIs
Calculates CTFU (Corrected Total Fluorescence Units)
"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import tifffile
from pathlib import Path
import json
import openpyxl
from openpyxl.styles import Font, PatternFill
import seaborn as sns
import argparse


class SegmentQuantifier:
    def __init__(self, roi_output_dir, channel_images=None):
        """Initialize quantifier."""
        self.roi_dir = Path(roi_output_dir)
        
        print(f"Loading ROI data from {self.roi_dir}")
        self.segments = np.load(self.roi_dir / 'segments.npy')
        
        with open(self.roi_dir / 'metadata.json', 'r') as f:
            self.metadata = json.load(f)
        
        self.n_segments = self.metadata['n_segments']
        self.downscale = self.metadata.get('downscale', 1.0)
        
        print(f"Loaded {self.n_segments} segments")
        print(f"Segment shape: {self.segments.shape}")
        
        self.channel_images = channel_images or {}
        self.channel_data = {}
        self.background_values = {}
        
    def add_channel(self, channel_name, image_path, background_roi=None, downscale=False):
        """
        Add a channel image for quantification.
        
        Parameters
        ----------
        channel_name : str
            Name of the channel/probe
        image_path : str or Path
            Path to channel TIFF image
        background_roi : array-like, optional
            Mask or coordinates for background region
        downscale : bool
            Whether to downscale (should be False for full-res quantification)
        """
        print(f"Loading channel: {channel_name}")
        image = tifffile.imread(image_path)
        print(f"  Image shape: {image.shape}")
        
        # Check shape matches segments
        if image.shape != self.segments.shape:
            raise ValueError(
                f"Channel image shape {image.shape} doesn't match segments {self.segments.shape}. "
                f"Make sure you're using full-resolution channel images!"
            )
        
        self.channel_data[channel_name] = image
        
        # Calculate background
        self.background_values[channel_name] = self.calculate_background_automatic(channel_name)
    
    def calculate_background_automatic(self, channel_name, percentile_range=(5, 20)):
        """
        Automatically calculate background from low-intensity tissue regions.
        
        Parameters
        ----------
        channel_name : str
            Which channel to calculate background for
        percentile_range : tuple
            (min, max) percentiles to use for background estimation
            Default: (5, 20) uses pixels between 5th-20th percentile of tissue
        """
        if channel_name not in self.channel_data:
            raise ValueError(f"Channel {channel_name} not loaded yet!")
        
        image = self.channel_data[channel_name]
        
        # Load epithelial mask (or any tissue mask) to restrict to tissue regions
        epithelial_mask_path = self.roi_dir / 'epithelial_mask.npy'
        if epithelial_mask_path.exists():
            tissue_mask = np.load(epithelial_mask_path)
            tissue_pixels = image[tissue_mask]
        else:
            # Fallback: use all non-zero pixels
            tissue_pixels = image[image > 0]
        
        # Get pixels in the bottom percentile range (low intensity tissue)
        low_percentile = np.percentile(tissue_pixels, percentile_range[0])
        high_percentile = np.percentile(tissue_pixels, percentile_range[1])
        
        # Background = mean of pixels in this range
        background_pixels = tissue_pixels[(tissue_pixels >= low_percentile) & 
                                        (tissue_pixels <= high_percentile)]
        
        background = background_pixels.mean()
        
        print(f"  Background (mean of {percentile_range[0]}-{percentile_range[1]}th percentile): {background:.2f}")
        print(f"    Based on {len(background_pixels)} tissue pixels")
        
        return background
    
    def quantify_segments(self):
        """
        Quantify all channels across all segments using CTFU.
        
        CTFU = IntDen - (Area × Mean Background)
        
        Returns
        -------
        pd.DataFrame
            Results with columns: Segment, Channel, Area, Mean, Min, Max, IntDen, RawIntDen, CTFU
        """
        print(f"\nQuantifying {len(self.channel_data)} channels across {self.n_segments} segments...")
        
        results = []
        
        for seg_num in range(1, self.n_segments + 1):
            seg_mask = self.segments == seg_num
            area = seg_mask.sum()
            
            if area == 0:
                print(f"  Warning: Segment {seg_num} has 0 pixels, skipping")
                continue
            
            for channel_name, channel_image in self.channel_data.items():
                # Extract pixel values in this segment
                pixel_values = channel_image[seg_mask]
                
                # Calculate metrics
                mean_intensity = pixel_values.mean()
                min_intensity = pixel_values.min()
                max_intensity = pixel_values.max()
                integrated_density = pixel_values.sum()
                raw_integrated_density = integrated_density
                
                # Background
                background = self.background_values.get(channel_name, 0)
                
                # CTFU = IntDen - (Area × Mean Background)
                ctfu = integrated_density - (area * background)
                
                results.append({
                    'Segment': seg_num,
                    'Channel': channel_name,
                    'Area': area,
                    'Mean': mean_intensity,
                    'Min': min_intensity,
                    'Max': max_intensity,
                    'IntDen': integrated_density,
                    'RawIntDen': raw_integrated_density,
                    'Background': background,
                    'CTFU': ctfu
                })
        
        df = pd.DataFrame(results)
        print(f"\nQuantification complete: {len(df)} measurements")
        return df
    
    def save_results_xlsx(self, df, output_path=None, sample_name=None):
        """
        Save results to Excel with nice formatting.
        
        Parameters
        ----------
        df : pd.DataFrame
            Results from quantify_segments()
        output_path : str or Path
            Where to save Excel file
        sample_name : str
            Name for the sheet
        """
        if output_path is None:
            output_path = self.roi_dir / 'quantification_results.xlsx'
        
        if sample_name is None:
            sample_name = self.roi_dir.name
        
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write each channel to its own sheet
            for channel in df['Channel'].unique():
                channel_df = df[df['Channel'] == channel].copy()
                channel_df = channel_df.drop('Channel', axis=1)
                
                sheet_name = f"{sample_name}_{channel}"[:31]  # Excel limit
                channel_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Format the sheet
                ws = writer.sheets[sheet_name]
                
                # Header formatting
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                
                # Adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
        
        print(f"Saved Excel results to {output_path}")
        return output_path
    
    def plot_results_publication(self, df, output_path=None):
        """
        Create publication-quality line plots.
        
        Parameters
        ----------
        df : pd.DataFrame
            Results from quantify_segments()
        output_path : str or Path
            Where to save plot
        """
        # Set publication style
        plt.style.use('seaborn-v0_8-paper')
        sns.set_palette("husl")
        
        # Create figure
        fig, ax = plt.subplots(figsize=(8, 5), dpi=300)
        
        # Get distinct colors
        channels = df['Channel'].unique()
        colors = sns.color_palette("husl", len(channels))
        
        # Plot each channel as a line
        for channel, color in zip(channels, colors):
            channel_data = df[df['Channel'] == channel].sort_values('Segment')
            
            ax.plot(
                channel_data['Segment'], 
                channel_data['CTFU'],
                marker='o',
                linewidth=2.5,
                markersize=8,
                label=channel,
                color=color,
                alpha=0.9
            )
        
        # Formatting
        ax.set_xlabel('Segment', fontsize=14, fontweight='bold', fontfamily='Arial')
        ax.set_ylabel('CTFU', fontsize=14, fontweight='bold', fontfamily='Arial')
        ax.set_title('Corrected Total Fluorescence by Segment', 
                    fontsize=16, fontweight='bold', fontfamily='Arial', pad=20)
        
        # Grid
        ax.grid(True, alpha=0.3, linestyle='--', linewidth=0.5)
        ax.set_axisbelow(True)
        
        # Legend
        ax.legend(
            loc='best',
            frameon=True,
            fancybox=True,
            shadow=True,
            fontsize=11,
            framealpha=0.95
        )
        
        # Tick parameters
        ax.tick_params(labelsize=11)
        ax.set_xticks(range(1, self.n_segments + 1))
        
        # Spine styling
        for spine in ax.spines.values():
            spine.set_linewidth(1.5)
        
        plt.tight_layout()
        
        if output_path:
            plt.savefig(output_path, dpi=300, bbox_inches='tight', facecolor='white')
            print(f"Saved publication plot to {output_path}")
        else:
            plt.show()
        
        return fig
    
    def run_quantification(self, channel_dict, output_prefix='quantification', sample_name=None):
        """
        Run full quantification pipeline.
        
        Parameters
        ----------
        channel_dict : dict
            Mapping of channel names to image paths
        output_prefix : str
            Prefix for output files
        sample_name : str
            Sample name for Excel sheets
        """
        # Load all channels
        for channel_name, image_path in channel_dict.items():
            self.add_channel(channel_name, image_path)
        
        # Quantify
        df = self.quantify_segments()
        
        # Save results
        xlsx_path = self.roi_dir / f'{output_prefix}_results.xlsx'
        self.save_results_xlsx(df, xlsx_path, sample_name)
        
        # Plot
        plot_path = self.roi_dir / f'{output_prefix}_plot.png'
        self.plot_results_publication(df, plot_path)
        
        print(f"\nQuantification complete!")
        print(f"  Excel: {xlsx_path}")
        print(f"  Plot: {plot_path}")
        
        return df


def select_file_gui(title="Select file", filetypes=[("TIFF files", "*.tif *.tiff")]):
    """Open GUI file picker."""
    from tkinter import Tk, filedialog
    
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    file_path = filedialog.askopenfilename(
        title=title,
        filetypes=filetypes
    )
    
    root.destroy()
    return file_path

def select_file_in_folder(folder_path, title="Select file"):
    """Open file picker restricted to specific folder."""
    from tkinter import Tk, filedialog
    
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    file_path = filedialog.askopenfilename(
        title=title,
        initialdir=str(folder_path),
        filetypes=[("TIFF files", "*.tif *.tiff"), ("All files", "*.*")]
    )
    
    root.destroy()
    return file_path


def main():
    """Main entry point with test mode support."""
    parser = argparse.ArgumentParser(
        description='Quantify segmented intestinal ROIs'
    )
    parser.add_argument('roi_dir', nargs='?', help='ROI output directory')
    parser.add_argument('--test', action='store_true',
                       help='Test mode: interactive file selection')
    parser.add_argument('--batch', action='store_true',
                        help='Batch mode: process multiple samples')
    
    args = parser.parse_args()

    if args.batch:
        print("\n" + "="*60)
        print("BATCH MODE: Multiple Sample Quantification")
        print("="*60)
        
        print("\nSelect parent directory containing ROI output folders...")
        from tkinter import Tk, filedialog
        root = Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        
        parent_dir = filedialog.askdirectory(title="Select parent directory with ROI outputs")
        root.destroy()
        
        if not parent_dir:
            print("No directory selected!")
            return
        
        # Find all ROI output directories (contain segments.npy)
        parent_path = Path(parent_dir)
        roi_dirs = [d for d in parent_path.iterdir() 
                    if d.is_dir() and (d / 'segments.npy').exists()]
        
        if not roi_dirs:
            print("No ROI output directories found!")
            return
        
        print(f"\nFound {len(roi_dirs)} samples:")
        for roi_dir in roi_dirs:
            print(f"  - {roi_dir.name}")
        
        # Ask how many channels (same for all samples)
        n_channels = int(input("\nHow many channels (same for all samples)? "))
        
        # Get channel names
        channel_names = []
        for i in range(n_channels):
            name = input(f"Channel {i+1} name: ").strip()
            channel_names.append(name)
        
        # Process each sample
        for i, roi_dir in enumerate(roi_dirs, 1):
            print("\n" + "="*60)
            print(f"QUANTIFYING SAMPLE {i}/{len(roi_dirs)}: {roi_dir.name}")
            print("="*60)
            
            try:
                quantifier = SegmentQuantifier(roi_dir)
                
                # Get sample name
                sample_name = input(f"\nSample name [default: {roi_dir.name}]: ").strip()
                if not sample_name:
                    sample_name = roi_dir.name
                
                # Find original image folder (go up from roi_output)
                metadata_path = roi_dir / 'metadata.json'
                with open(metadata_path, 'r') as f:
                    metadata = json.load(f)
                
                original_image_path = Path(metadata['image_path'])
                image_folder = original_image_path.parent
                
                # Select channels in that folder
                channel_dict = {}
                for channel_name in channel_names:
                    print(f"\nSelect {channel_name} image in {image_folder.name}...")
                    channel_path = select_file_in_folder(
                        image_folder,
                        f"Select {channel_name} in {image_folder.name}"
                    )
                    
                    if channel_path:
                        print(f"  Selected: {Path(channel_path).name}")
                        channel_dict[channel_name] = channel_path
                    else:
                        print(f"  No file selected for {channel_name}, skipping")
                
                if channel_dict:
                    quantifier.run_quantification(channel_dict, sample_name=sample_name)
                else:
                    print(f"No channels selected for {sample_name}, skipping")
                    
            except Exception as e:
                print(f"\nERROR quantifying {roi_dir.name}: {e}")
        
        print("\n" + "="*60)
        print("BATCH QUANTIFICATION COMPLETE")
        print("="*60)
    
    # Test mode: interactive selection
    elif args.test:
        print("\n" + "="*60)
        print("TEST MODE: Interactive Quantification")
        print("="*60)
        
        # Find ROI directory
        from tkinter import Tk, filedialog
        root = Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        
        print("\nSelect ROI output directory...")
        roi_dir = filedialog.askdirectory(title="Select ROI output directory")
        root.destroy()
        
        if not roi_dir:
            print("No directory selected!")
            return
        
        print(f"Selected: {roi_dir}")
        
        # Initialize quantifier
        quantifier = SegmentQuantifier(roi_dir)
        
        # Get sample name
        sample_name = input("\nSample name: ").strip()
        if not sample_name:
            sample_name = Path(roi_dir).name
        
        # Select channels
        n_channels = int(input("\nHow many channels to quantify? "))
        
        channel_dict = {}
        for i in range(n_channels):
            print(f"\n--- Channel {i+1} ---")
            channel_name = input("Channel name: ").strip()
            
            print(f"Select {channel_name} image...")
            channel_path = select_file_gui(f"Select {channel_name} image")
            
            if not channel_path:
                print(f"No file selected for {channel_name}, skipping")
                continue
            
            print(f"  Selected: {Path(channel_path).name}")
            channel_dict[channel_name] = channel_path
        
        if len(channel_dict) == 0:
            print("\nNo channels added!")
            return
        
        # Run quantification
        print(f"\nQuantifying {len(channel_dict)} channels...")
        quantifier.run_quantification(channel_dict, sample_name=sample_name)
        
    # Batch mode: command line
    elif args.roi_dir:
        quantifier = SegmentQuantifier(args.roi_dir)
        
        # Interactive channel addition
        print("\n" + "="*60)
        print("ADD CHANNELS FOR QUANTIFICATION")
        print("="*60)
        
        channel_dict = {}
        while True:
            channel_name = input("\nEnter channel name (or 'done' to finish): ").strip()
            if channel_name.lower() == 'done':
                break
            
            image_path = input(f"Enter path to {channel_name} image: ").strip()
            channel_dict[channel_name] = image_path
        
        if len(channel_dict) == 0:
            print("No channels added!")
            return
        
        sample_name = input("\nEnter sample name: ").strip()
        quantifier.run_quantification(channel_dict, sample_name=sample_name)
        
    else:
        parser.print_help()


if __name__ == "__main__":
    main()